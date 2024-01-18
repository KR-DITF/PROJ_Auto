import openpyxl
from openpyxl import workbook, load_workbook

wb = load_workbook("saraksts_IMEI_konti.xlsx") #atveram galveno workbook, kurā atrodas visi konta nr. un to statusi
ws = wb.active
ws.title = "konti"
max_r = ws.max_row 

konti = []
for row1 in range (2,max_r+1): #ejam cauri visiem konta nr.
    status = ws["C"+str(row1)].value
    if status == None: #pārbaudam vai kontam ir vai nav statuss, ja nav, tad pievienojam kontu sarakstā
        k_nr = ws["A"+str(row1)].value
        konti.append(k_nr) #pievienojam kontu sarakstam
print(len(konti))
wb1 = load_workbook("aktualie.xlsx") #atver otro workbook, no kuras nolasa vai konts atrodas sarakstā un vai bilance <=10
ws1 = wb1.active
ws1.title = "Lapa1"
max_ro = ws1.max_row
parb_k = []
for sk in konti: #sāk loop, lai izietu cauri kontu nr., kuri vēl nav apstrādāti
    flag = True
    for row2 in range(2,max_ro+1): #katram kontam iet cauri otrajam workbook
        ch = ws1["A"+str(row2)].value
        bal = ws1["E"+str(row2)].value

        if sk == ch and bal > 10: #pārbauda vai konta nr. atrodas failā un bilanse ir lielāka par 10
            flag = False #ja konts atrodas listā un bilanse > 10, nomaina pārbaudi uz True
            break
        elif sk == ch and bal <= 10: #ja konts nav sarakstā, vai tam bilanse ir mazāka par 10, pievieno sarakstam, kurus tālāk jāpārbauda
            flag = False
            parb_k.append(sk)
            break
    if flag:
        parb_k.append(sk)
print(len(parb_k))
sheets = wb.sheetnames
output = []
output1 = []
for ko_nr in parb_k: #iet cauri visiem atrastajiem kontiem
    flag1 = False
    n = 4 #sāk ar 5. sheet
    while n < len(sheets): #excel dokuments mainās katru nedēļu, tapēc jāatrod cik daudz lapas ir dokumentā
        ws = wb[sheets[n]]
        max_row = ws.max_row
        for row3 in range(2,max_row+1): #iet cauri sheeta visām rindām un pārbauda vai iegūtie kodi sakrīt ar dotajiem
            tempk = ws["A"+str(row3)].value
            imei = ws["C"+str(row3)].value
            if ko_nr == tempk and imei != None: #pārbauda vai iegūtais kods sakrīt ar dotajiem
                msisdn = ws["B"+str(row3)].value
                imei2 = ws["D"+str(row3)].value
                output1.append(ko_nr) #ja sakrīt, sarakstam pievieno konta nr., msisdn un abus imei kodus, ja tādi ir vairāki
                output1.append(msisdn)
                output1.append(imei)
                output1.append(imei2)
                for g in output1:
                    output.append(g)
                output1.clear()
                flag1 = True #ja tika atrasts kods, nomaina pārbaudi uz patiesi
        if flag1: #ja tik atrasts kods, tad pāriet uz nākamo kodu, kuru vajag pārbaudīt
            break
        else: #ja neatrada kodu, ver vaļā nākamo sheet
            n+=1

print(output,len(output))        

wb.close()
wb1.close()

wb2 = load_workbook("finals.xlsx") #atver rezultātu excel
ws2 = wb2.active
max_row2 = ws2.max_row
ws2.delete_row(1,max_row2) #iztīra visu lapu, lai būtu jaunākie dati

inp = 2
a = 0
ws2["A1"].value = "Konta nr."
ws2["B1"].value = "msisdn"
ws2["C1"].value = "IMEI"
ws2["D1"].value = "IMEI 2"

while a < (len(output)-4): #ievieto iegūtos datus excelī
    ws2["A"+str(inp)].value = output[a]
    ws2["B"+str(inp)].value = output[a+1]
    ws2["C"+str(inp)].value = output[a+2]
    ws2["D"+str(inp)].value = output[a+3]
    a+=4
    inp+=1

wb2.save("konti.xlsx") #saglabā datus
wb2.close()
